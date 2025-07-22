from flask import Blueprint, render_template, request, jsonify, session, send_file
import os
from werkzeug.utils import secure_filename
from app.utils.sbom_analyzer import load_dummy_sbom
from app.utils.vulnerability_database import VulnerabilityDatabase
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

main = Blueprint('main', __name__)

# Updated to accept project dependency files instead of SBOM files
ALLOWED_EXTENSIONS = {
    'txt',      # requirements.txt
    'toml',     # pyproject.toml, Pipfile
    'json',     # package.json, package-lock.json
    'xml',      # pom.xml
    'lock',     # Cargo.lock, yarn.lock
    'gradle',   # build.gradle
    'mod',      # go.mod
    'sum',      # go.sum
    'zip'       # full project uploads
}

def allowed_file(filename):
    """Allow only requirements.txt, Pipfile, or pyproject.toml (case-insensitive)."""
    filename_lower = filename.lower()
    allowed_names = [
        'requirements.txt',
        'pipfile',
        'pyproject.toml',
    ]
    return any(filename_lower.endswith(name) for name in allowed_names)

def detect_project_type(filename):
    """Detect project type based on filename"""
    filename_lower = filename.lower()
    if 'requirements.txt' in filename_lower or filename_lower.endswith('.txt'):
        return 'python'
    elif 'pyproject.toml' in filename_lower:
        return 'python'
    elif filename == 'pipfile' or 'pipfile' in filename_lower:
        return 'python'
    elif 'package.json' in filename_lower or 'yarn.lock' in filename_lower:
        return 'node'
    elif 'pom.xml' in filename_lower or filename_lower.endswith('.gradle'):
        return 'java'
    elif 'go.mod' in filename_lower or 'go.sum' in filename_lower:
        return 'go'
    elif 'cargo.toml' in filename_lower or 'cargo.lock' in filename_lower:
        return 'rust'
    elif filename_lower.endswith('.zip'):
        return 'archive'
    else:
        return 'unknown'

@main.route('/')
def index():
    """Main page for SBOM generation and analysis"""
    return render_template('index.html')

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')

@main.route('/upload', methods=['POST'])
def upload_file():
    """Handle project file upload and generate SBOM"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(str(file.filename))
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            file.save(filepath)

            project_type = detect_project_type(filename)
            sbom_data = load_dummy_sbom(filepath)

            # Use dummy data directly
            scan_id = str(uuid.uuid4())
            db = VulnerabilityDatabase()
            db.save_scan_result(
                scan_id=scan_id,
                dependencies=sbom_data.get('dependencies', []),
                subdependencies=sbom_data.get('dependency_tree', []),
                vulnerabilities=sbom_data.get('vulnerabilities', {}),
                skipped_dependencies=sbom_data.get('skipped_dependencies', []),
                user_id=None,
                filename=filename,
                package_health_data=sbom_data.get('package_health_issues', {})
            )

            session['scan_id'] = scan_id
            session['project_type'] = project_type
            session['filename'] = filename
            session['package_health_data'] = sbom_data.get('package_health_issues', {})

            return jsonify({
                'message': 'SBOM generated successfully',
                'project_type': project_type,
                'filename': filename,
                'scan_id': scan_id,
                'skipped_dependencies': sbom_data.get('skipped_dependencies', []),
                'redirect_url': '/results'
            })
        except Exception as e:
            return jsonify({'error': f'SBOM generation failed: {str(e)}'}), 500
    return jsonify({'error': 'Invalid file type. Please upload requirements.txt, pyproject.toml, or Pipfile.'}), 400

@main.route('/results')
def results():
    """Display SBOM analysis results"""
    return render_template('results.html')

@main.route('/api/health')
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'service': 'sbom-generator'})

@main.route('/api/vulnerabilities/<package_name>')
def get_package_vulnerabilities(package_name):
    """Get vulnerability data for a specific package"""
    try:
        scan_id = session.get('scan_id')
        if not scan_id:
            return jsonify({'error': 'No scan ID found in session.'}), 400
        db = VulnerabilityDatabase()
        scan_result = db.get_scan_result_by_id(scan_id)
        if not scan_result:
            return jsonify({'error': 'No scan result found for this scan ID.'}), 404
        vulnerability_data = scan_result.get('vulnerabilities', {})
        package_risk = vulnerability_data.get(package_name)
        if package_risk:
            return jsonify(package_risk)
        # Return safe default if package not found
        return jsonify({
            'package_name': package_name,
            'version': 'unknown',
            'risk_score': 0.0,
            'risk_level': 'safe',
            'vulnerability_count': 0,
            'vulnerabilities': [],
            'recommendation': 'No vulnerabilities found for this package.'
        })
    except Exception as e:
        print(f"Error in vulnerability API: {e}")
        return jsonify({
            'package_name': package_name,
            'version': 'unknown',
            'risk_score': 0.0,
            'risk_level': 'safe',
            'vulnerability_count': 0,
            'vulnerabilities': [],
            'recommendation': 'No vulnerabilities found for this package.'
        }) 

@main.route('/api/package_health')
def get_package_health():
    """Get package health data for the current session or scan_id parameter"""
    try:
        # Try to get scan_id from session first
        scan_id = session.get('scan_id')
        
        # If not in session, try to get from request parameters
        if not scan_id:
            scan_id = request.args.get('scan_id')
        
        if not scan_id:
            return jsonify({'error': 'No scan ID found in session or parameters.'}), 400
        
        db = VulnerabilityDatabase()
        scan_result = db.get_scan_result_by_id(scan_id)
        if not scan_result:
            return jsonify({'error': 'No scan result found for this scan ID.'}), 404
        
        package_health_data = scan_result.get('package_health_issues', {})
        return jsonify(package_health_data)
    except Exception as e:
        print(f"Error getting package health data: {e}")
        return jsonify({'error': 'Failed to retrieve package health data.'}), 500

@main.route('/api/scan_data')
def get_scan_data():
    """Return all scan data for the current scan_id in the session."""
    scan_id = session.get('scan_id')
    project_type = session.get('project_type')
    filename = session.get('filename')
    if not scan_id:
        return jsonify({'error': 'No scan ID found in session.'}), 400
    db = VulnerabilityDatabase()
    scan_result = db.get_scan_result_by_id(scan_id)
    if not scan_result:
        return jsonify({'error': 'No scan result found for this scan ID.'}), 404
    
    # Defensive: ensure dependencies and subdependencies are lists
    dependencies = scan_result.get('dependencies', [])
    if not isinstance(dependencies, list):
        dependencies = []
    subdependencies = scan_result.get('subdependencies', [])
    if not isinstance(subdependencies, list):
        subdependencies = []
    # Defensive: ensure vulnerabilities is a dict
    vulnerabilities = scan_result.get('vulnerabilities', {})
    if not isinstance(vulnerabilities, dict):
        vulnerabilities = {}
    # Defensive: ensure project_type and filename are strings
    project_type = project_type if isinstance(project_type, str) else 'Unknown'
    filename = filename if isinstance(filename, str) else '-'
    skipped_dependencies = scan_result.get('skipped_dependencies', [])
    
    # Debug output for troubleshooting
    # print('=== /api/scan_data DEBUG ===')
    # print('Dependencies:', dependencies)
    # print('Vulnerabilities:', vulnerabilities)
    # print('Skipped dependencies:', skipped_dependencies)
    # print('Project type:', project_type)
    # print('Filename:', filename)
    # print('-----------------------------')
    return jsonify({
        'dependencies': dependencies,
        'dependency_tree': subdependencies,
        'vulnerabilities': vulnerabilities,
        'project_type': project_type,
        'filename': filename,
        'skipped_dependencies': skipped_dependencies,
        'sbom_data': {
            'dependencies': dependencies,
            'dependency_tree': subdependencies
        }
    }) 

@main.route('/api/recent_scans')
def api_recent_scans():
    """Return a list of recent scan summaries from the last 24 hours."""
    try:
        db = VulnerabilityDatabase()
        scans = db.get_recent_scans(hours=24)
        return jsonify({'recent_scans': scans})
    except Exception as e:
        return jsonify({'error': f'Failed to fetch recent scans: {str(e)}'}), 500

@main.route('/api/scan_result/<scan_id>')
def api_scan_result(scan_id):
    """Return the full scan result for a given scan_id."""
    try:
        db = VulnerabilityDatabase()
        scan_result = db.get_scan_result_by_id(scan_id)
        if not scan_result:
            return jsonify({'error': 'No scan result found for this scan ID.'}), 404
        return jsonify({'scan_result': scan_result})
    except Exception as e:
        return jsonify({'error': f'Failed to fetch scan result: {str(e)}'}), 500 

@main.route('/set_scan_id', methods=['POST'])
def set_scan_id():
    data = request.get_json()
    scan_id = data.get('scan_id')
    if not scan_id:
        return jsonify({'error': 'No scan_id provided'}), 400
    session['scan_id'] = scan_id
    return jsonify({'success': True}) 

@main.route('/clear_old_scans', methods=['POST'])
def clear_old_scans():
    """Clear all scan results from the database, but preserve GitHub enrichment data for reuse."""
    try:
        db = VulnerabilityDatabase()
        db.clear_all_scan_results()
        return jsonify({'success': True, 'message': 'Scan results cleared (GitHub data preserved)'})
    except Exception as e:
        return jsonify({'error': f'Failed to clear scans: {str(e)}'}), 500

@main.route('/clear_cache', methods=['POST'])
def clear_cache():
    """Completely clear all cache and recreate database."""
    try:
        db = VulnerabilityDatabase()
        db.clear_all_data()
        return jsonify({'success': True, 'message': 'Database completely cleared and recreated'})
    except Exception as e:
        return jsonify({'error': f'Failed to clear cache: {str(e)}'}), 500

@main.route('/export_excel', methods=['POST'])
def export_excel():
    """Export SBOM analysis results to Excel file"""
    try:
        scan_id = session.get('scan_id')
        project_type = session.get('project_type', 'Unknown')
        filename = session.get('filename', 'Unknown')
        
        if not scan_id:
            return jsonify({'error': 'No scan ID found in session.'}), 400
            
        db = VulnerabilityDatabase()
        scan_result = db.get_scan_result_by_id(scan_id)
        if not scan_result:
            return jsonify({'error': 'No scan result found for this scan ID.'}), 404
            
        dependencies = scan_result.get('dependencies', [])
        subdependencies = scan_result.get('subdependencies', [])
        vulnerability_data = scan_result.get('vulnerabilities', {})
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # Summary sheet
        summary_sheet = wb.create_sheet('Summary')
        
        # Add title
        title_cell = summary_sheet['A1']
        title_cell.value = 'SBOM Analysis Report'
        title_cell.font = Font(size=16, bold=True)
        summary_sheet.merge_cells('A1:C1')
        
        # Project information
        summary_sheet['A3'] = 'Project Information'
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['A4'] = 'Project Type'
        summary_sheet['B4'] = project_type
        summary_sheet['A5'] = 'File Analyzed'
        summary_sheet['B5'] = filename
        summary_sheet['A6'] = 'Analysis Date'
        summary_sheet['B6'] = scan_result.get('timestamp', 'Unknown')
        
        # Summary statistics
        summary_sheet['A8'] = 'Summary Statistics'
        summary_sheet['A8'].font = Font(bold=True)
        summary_sheet['A9'] = 'Direct Dependencies'
        summary_sheet['B9'] = len(dependencies)
        summary_sheet['A10'] = 'Total Dependencies'
        
        # Count total dependencies including subdependencies
        def count_total_deps(dep_list):
            total = 0
            for dep in dep_list:
                total += 1  # Count the dependency itself
                if dep.get('subdependencies'):
                    total += count_total_deps(dep['subdependencies'])
            return total
        
        total_deps = count_total_deps(subdependencies) if subdependencies else len(dependencies)
        summary_sheet['B10'] = total_deps
        
        # Risk distribution
        summary_sheet['A12'] = 'Risk Distribution'
        summary_sheet['A12'].font = Font(bold=True)
        
        risk_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'safe': 0}
        for dep in dependencies:
            vuln_data = vulnerability_data.get(dep.get('name', ''))
            if vuln_data:
                risk_level = vuln_data.get('risk_level', 'safe').lower()
                if risk_level in risk_counts:
                    risk_counts[risk_level] += 1
            else:
                risk_counts['safe'] += 1
        
        summary_sheet['A13'] = 'Critical Risk'
        summary_sheet['B13'] = risk_counts['critical']
        summary_sheet['A14'] = 'High Risk'
        summary_sheet['B14'] = risk_counts['high']
        summary_sheet['A15'] = 'Medium Risk'
        summary_sheet['B15'] = risk_counts['medium']
        summary_sheet['A16'] = 'Low Risk'
        summary_sheet['B16'] = risk_counts['low']
        summary_sheet['A17'] = 'Safe'
        summary_sheet['B17'] = risk_counts['safe']
        
        # Dependencies sheet
        deps_sheet = wb.create_sheet('Dependencies')
        
        # Headers
        headers = ['Package Name', 'Version', 'Type', 'Risk Level', 'Vulnerability Count', 'Risk Score', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = deps_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Collect all dependencies including subdependencies, with type
        all_deps = []
        
        def collect_deps(dep_list, level=0, parent_is_direct=True):
            for dep in dep_list:
                # Add prefix to show hierarchy
                prefix = "  " * level
                dep_name = f"{prefix}{dep.get('name', 'Unknown')}"
                dep_type = 'Direct' if level == 0 and parent_is_direct else 'Sub'
                all_deps.append({
                    'name': dep_name,
                    'clean_name': dep.get('name', 'Unknown'),
                    'version': dep.get('version', 'Unknown'),
                    'type': dep_type,
                    'level': level
                })
                # Recursively add subdependencies
                if dep.get('subdependencies'):
                    collect_deps(dep['subdependencies'], level + 1, parent_is_direct=False)
        
        # Collect all dependencies from subdependencies (which contains the full tree)
        collect_deps(subdependencies)
        
        # Data
        for row, dep in enumerate(all_deps, 2):
            dep_name = dep['name']
            clean_name = dep['clean_name']
            vuln_data = vulnerability_data.get(clean_name, {})
            deps_sheet.cell(row=row, column=1).value = dep_name
            deps_sheet.cell(row=row, column=2).value = dep['version']
            deps_sheet.cell(row=row, column=3).value = dep['type']
            deps_sheet.cell(row=row, column=4).value = vuln_data.get('risk_level', 'SAFE').upper()
            deps_sheet.cell(row=row, column=5).value = vuln_data.get('vulnerability_count', 0)
            deps_sheet.cell(row=row, column=6).value = vuln_data.get('risk_score', 0.0)
            deps_sheet.cell(row=row, column=7).value = vuln_data.get('recommendation', 'No vulnerabilities found')
        
        # Vulnerabilities sheet (if any exist)
        all_vulnerabilities = []
        for pkg_name, pkg_data in vulnerability_data.items():
            if pkg_data.get('vulnerabilities'):
                for vuln in pkg_data['vulnerabilities']:
                    # Get package version
                    pkg_version = pkg_data.get('version', 'Unknown')
                    
                    # Get CVE ID (prefer cve_id, fallback to id)
                    cve_id = vuln.get('cve_id', '') or vuln.get('id', '')
                    
                    # Get risk level from package data
                    risk_level = pkg_data.get('risk_level', 'SAFE').upper()
                    
                    # Get severity from vulnerability
                    severity = vuln.get('severity', 'UNKNOWN').upper()
                    
                    # Get summary (use title or first 100 chars of description)
                    summary = vuln.get('title', '')
                    if not summary and vuln.get('description'):
                        summary = vuln.get('description', '')[:100] + ('...' if len(vuln.get('description', '')) > 100 else '')
                    
                    # Get references
                    references = '; '.join(vuln.get('references', []))
                    
                    all_vulnerabilities.append({
                        'package': pkg_name,
                        'version': pkg_version,
                        'risk_level': risk_level,
                        'cve_id': cve_id,
                        'severity': severity,
                        'summary': summary,
                        'references': references
                    })
        
        if all_vulnerabilities:
            vuln_sheet = wb.create_sheet('Vulnerabilities')
            
            # Add title
            title_cell = vuln_sheet['A1']
            title_cell.value = 'Vulnerability Analysis Report'
            title_cell.font = Font(size=14, bold=True)
            vuln_sheet.merge_cells('A1:G1')
            
            # Headers (matching your table format)
            vuln_headers = ['Package', 'Version', 'Risk Level', 'CVE ID(s)', 'Severity', 'Summary', 'References']
            for col, header in enumerate(vuln_headers, 1):
                cell = vuln_sheet.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row, vuln in enumerate(all_vulnerabilities, 4):
                vuln_sheet.cell(row=row, column=1).value = vuln['package']
                vuln_sheet.cell(row=row, column=2).value = vuln['version']
                vuln_sheet.cell(row=row, column=3).value = vuln['risk_level']
                vuln_sheet.cell(row=row, column=4).value = vuln['cve_id']
                vuln_sheet.cell(row=row, column=5).value = vuln['severity']
                vuln_sheet.cell(row=row, column=6).value = vuln['summary']
                vuln_sheet.cell(row=row, column=7).value = vuln['references']
                
                # Color code severity levels
                severity_cell = vuln_sheet.cell(row=row, column=5)
                severity = vuln['severity'].lower()
                if severity in ['critical', 'high']:
                    severity_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif severity == 'medium':
                    severity_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif severity == 'low':
                    severity_cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        # Set reasonable column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if sheet == 'Vulnerabilities':
                ws.column_dimensions['A'].width = 25  # Package
                ws.column_dimensions['B'].width = 12  # Version
                ws.column_dimensions['C'].width = 12  # Risk Level
                ws.column_dimensions['D'].width = 20  # CVE ID(s)
                ws.column_dimensions['E'].width = 12  # Severity
                ws.column_dimensions['F'].width = 50  # Summary
                ws.column_dimensions['G'].width = 60  # References
            else:
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 30
                ws.column_dimensions['G'].width = 40
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'sbom_analysis_{scan_id[:8]}.xlsx'
        )
        
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        return jsonify({'error': f'Failed to export Excel: {str(e)}'}), 500 

@main.route('/export_spdx', methods=['POST'])
def export_spdx():
    """Export SBOM in SPDX format as a downloadable JSON file"""
    try:
        scan_id = session.get('scan_id')
        project_type = session.get('project_type', 'Unknown')
        filename = session.get('filename', 'Unknown')
        if not scan_id:
            return jsonify({'error': 'No scan ID found in session.'}), 400
        db = VulnerabilityDatabase()
        scan_result = db.get_scan_result_by_id(scan_id)
        if not scan_result:
            return jsonify({'error': 'No scan result found for this scan ID.'}), 404
        # SPDX SBOM is generated at upload and stored in sbom_data['sbom_data']
        # If not present, regenerate from dependencies
        spdx_data = None
        if 'sbom_data' in scan_result:
            spdx_data = scan_result['sbom_data']
        elif 'dependencies' in scan_result:
            from app.utils.sbom_analyzer import generate_spdx_sbom
            spdx_data = generate_spdx_sbom(scan_result['dependencies'], project_type, filename)
        else:
            return jsonify({'error': 'No SBOM data found for this scan.'}), 500
        from io import BytesIO
        import json
        spdx_bytes = BytesIO(json.dumps(spdx_data, indent=2).encode('utf-8'))
        spdx_bytes.seek(0)
        return send_file(
            spdx_bytes,
            mimetype='application/json',
            as_attachment=True,
            download_name=f'sbom_{scan_id[:8]}.spdx.json'
        )
    except Exception as e:
        print(f"Error exporting SPDX: {e}")
        return jsonify({'error': f'Failed to export SPDX: {str(e)}'}), 500 